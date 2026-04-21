"""Sync market-brief-2026-04-21 v2 data to master-data-log.xlsx."""
from openpyxl import load_workbook
from datetime import date, datetime

today = date(2026, 4, 21)
source_brief = "2026-04-21/market-brief-2026-04-21.md v2"

wb = load_workbook("master-data-log.xlsx")


def find_or_append_row(ws, today):
    for r in range(2, ws.max_row + 1):
        d = ws.cell(r, 1).value
        if isinstance(d, datetime):
            d = d.date()
        if d == today:
            return r
    return ws.max_row + 1


def write_row(ws, row_num, headers, data):
    for col, h in enumerate(headers, 1):
        ws.cell(row_num, col, data.get(h))


# ── DailyVariables ───────────────────────────────────────────────────────────
ws_dv = wb["DailyVariables"]
headers_dv = [ws_dv.cell(1, c).value for c in range(1, ws_dv.max_column + 1)]
target = find_or_append_row(ws_dv, today)

dv_data = {
    "Date": today,
    "VIX": 18.86,
    "VIX3M": 21.24,
    "VIX/VIX3M": round(18.86 / 21.24, 4),
    "MOVE": 67.9035,
    "DXY": 98.191,
    "HY_OAS": 2.85,
    "NFCI": -0.47,
    "Intermediary_Cap_Ratio": 0.1295,
    "Intermediary_Cap_Z": 1.65,
    "DGS2": round(4.25 - 0.54, 3),
    "DGS10": 4.25,
    "2s10s": 0.54,
    "DFII10": 1.89,
    "T10YIE": 2.36,
    "ACM_TP_10Y": None,
    "SPX": 7109.14,
    "NDX": 26590.344,
    "SPY": 708.72,
    "QQQ": 646.79,
    "EWJ": 89.33,
    "EWY": 150.09,
    "NVDA": 202.06,
    "TSLA": 392.50,
    "TSM": 366.24,
    "AAPL": 273.05,
    "GOOGL": 337.42,
    "AMZN": 248.28,
    "META": 670.91,
    "INTC": 65.70,
    "MU": 448.42,
    "PYPL": 51.46,
    "PLTR": 145.89,
    "WDC": 374.11,
    "ResidMom_NVDA": -1.00,
    "ResidMom_TSLA": -9.12,
    "ResidMom_TSM": 1.00,
    "ResidMom_AAPL": 4.68,
    "ResidMom_GOOGL": -0.80,
    "ResidMom_AMZN": 1.39,
    "ResidMom_META": 0.75,
    "ResidMom_INTC": 13.89,
    "ResidMom_MU": 15.63,
    "ResidMom_PYPL": -5.25,
    "ResidMom_PLTR": -35.62,
    "ResidMom_WDC": 2.38,
    "Brent": 90.73,
    "WTI": 87.55,
    "Gold": 4812.0,
    "Silver": 79.19,
    "Copper": 6.06,
    "Platinum": 2096.2,
    "Palladium": 1576.5,
    "BasisMom_Brent_4w": 0.95,
    "BasisMom_Brent_12w": 2.95,
    "BasisMom_WTI_4w": 6.98,
    "BasisMom_WTI_12w": 6.98,
    "BasisMom_Gold_4w": -20.0,
    "BasisMom_Gold_12w": -20.0,
    "BasisMom_Silver_4w": -0.29,
    "BasisMom_Silver_12w": -0.29,
    "BasisMom_Copper_4w": 0.09,
    "BasisMom_Copper_12w": 0.09,
    "EURUSD": 1.1767,
    "USDJPY": 159.095,
    "BTC": 76374.86,
    "ETH": 2316.71,
    "BTC_HashRate": 876.26,
    "BTC_ActiveAddr": 464309.0,
    "BTC_ExchNetflow": None,
    "BTC_ETF_Flow": 412.0,
    "BTC_PerpFunding": None,
    "BTC_3mBasis": None,
    "ETH_ETF_Flow": None,
    "ETH_DailyTxns": None,
    "GradeA_Missing_Count": 1,
    "Source_Brief": source_brief,
}
write_row(ws_dv, target, headers_dv, dv_data)
print(f"DailyVariables: row {target} written for {today}")


# ── RegimeHistory ─────────────────────────────────────────────────────────────
ws_rh = wb["RegimeHistory"]
headers_rh = [ws_rh.cell(1, c).value for c in range(1, ws_rh.max_column + 1)]
rh_row = find_or_append_row(ws_rh, today)

rh_data = {
    "Date": today,
    "Regime_Label": "RISK-ON / GEOPOLITICAL-BINARY — IRAN CEASEFIRE EXPIRED (APR-22 UTC+8)",
    "Growth": "NEUTRAL-TO-POSITIVE",
    "Inflation": "MODERATING-TO-UNCERTAIN",
    "Policy": "DOVISH-LEAN-FADING",
    "FinConditions": "LOOSE-TIGHTENING-AT-MARGIN",
    "RiskOnOff": "RISK-ON-FRAGILE",
    "BTC_VolRegime": "SUPPORT-HOLD",
    "BTC_FundingBasis": "NEGATIVE-EXTREME",
    "BTC_Structural": "POSITIVE-addr464k>400k",
    "Watch_Var_1": "WTI $88/$95 (A) ceasefire binary",
    "Watch_Var_2": "VIX 17/20 (B) at pivot 18.86",
    "Watch_Var_3": "BTC $75K/$71.2K (A) 4th support test",
    "Promoted_Signals": "INTC+4,GOOGL+3,AAPL+3,SPY+3(P009),EWJ+3(P010),BTC+3(gate-blocked)",
    "NearMiss_Count": 5,
    "Source_Brief": source_brief,
}
write_row(ws_rh, rh_row, headers_rh, rh_data)
print(f"RegimeHistory: row {rh_row} written")


# ── DataQuality ───────────────────────────────────────────────────────────────
ws_dq = wb["DataQuality"]
headers_dq = [ws_dq.cell(1, c).value for c in range(1, ws_dq.max_column + 1)]
dq_row = find_or_append_row(ws_dq, today)

dq_data = {
    "Date": today,
    "Total_GradeA_Vars": 25,
    "Missing_Count": 1,
    "Missing_Rate_Pct": round(1 / 25 * 100, 1),
    "Missing_Variables": "GSCI (V034 commodity gate)",
    "Stale_Count": 4,
    "Stale_Variables": "BasisMom WTI/Gold/Silver/Copper T3 Apr-17 within 5d window",
    "Audit_ResidMom_Status": "LIVE",
    "Audit_IntCap_Status": "LIVE",
    "Audit_BasisMom_Status": "PARTIAL Brent-LIVE others-T3",
    "Pipeline_Health": "OPERATIONAL",
    "Source_Brief": source_brief,
    "Notes": "v2 T1 refresh: VIX3M/SPX/NDX/ETH/EWJ/EWY/USDJPY now LIVE. BTC ActiveAddr 464k>400k S flip. WTI $87.55 pre-mkt.",
    "T1_Count": 36,
    "T2_Count": 0,
    "T3_Cache_Count": 11,
    "T4_Missing_Count": 1,
    "Retrieval_Time_Sec": None,
    "Cache_Coverage_Pct": round(36 / 47 * 100, 1),
}
write_row(ws_dq, dq_row, headers_dq, dq_data)
print(f"DataQuality: row {dq_row} written")


# ── CatalystLog ───────────────────────────────────────────────────────────────
ws_cl = wb["CatalystLog"]
headers_cl = [ws_cl.cell(1, c).value for c in range(1, ws_cl.max_column + 1)]

existing_keys = set()
for r in range(2, ws_cl.max_row + 1):
    ev = (ws_cl.cell(r, 3).value or "")[:40]
    dt = str(ws_cl.cell(r, 2).value)[:10]
    existing_keys.add((dt, ev))

new_cats = [
    (date(2026, 4, 22), "GOOGL Q1 earnings (AC) GOOGL Sum+3 entry trigger",
     "GOOGL, NVDA, QQQ", "HIGH",
     "C+1 if beat (88% hist rate); entry after confirmed beat Apr-23; NVDA read-through"),
    (date(2026, 4, 22), "TSLA Q1 earnings (AC)",
     "TSLA", "MEDIUM",
     "No position; watch for short setup if miss; residual T=-1 active"),
    (date(2026, 4, 23), "INTC Q1 earnings (AC) Terafab validation",
     "INTC, QQQ", "HIGH",
     "C+1 if Terafab milestones hit; INTC Sum+4; trade-rec decides pre/post entry"),
    (date(2026, 4, 23), "AMZN Q1 earnings (AC)",
     "AMZN, QQQ", "MEDIUM",
     "Beat + cloud read-through could push AMZN to Sum+3"),
    (date(2026, 4, 28), "FOMC statement and dots",
     "All assets", "HIGH",
     "Primary SPY P009 invalidation; hawkish >2 cuts removed = exit if SPY<$700"),
    (date(2026, 4, 29), "META Q1 earnings same day FOMC",
     "META, QQQ", "MEDIUM",
     "Sum+2 currently; beat needed; FOMC same-day complicates C scoring"),
    (date(2026, 5, 12), "April CPI release",
     "Rates, DXY, Gold, BTC", "HIGH",
     "Inflation follow-through post Hormuz shock; key for rate path and commodity sleeve gate"),
    (date(2026, 5, 14), "Xi-Trump summit Beijing",
     "US-China tariffs, DXY, EWY", "HIGH",
     "Trade deal potential; tariff de-escalation C+1 for EWY/EWJ if confirmed"),
]

added = 0
for cat_date, event, assets, impact, expectation in new_cats:
    key = (str(cat_date)[:10], event[:40])
    if key not in existing_keys:
        nr = ws_cl.max_row + 1
        vals = {
            "Date_Added": today,
            "Catalyst_Date": cat_date,
            "Event": event,
            "Assets_Affected": assets,
            "Impact_Level": impact,
            "Expectation": expectation,
            "Outcome": "PENDING",
            "Source_File": source_brief,
        }
        write_row(ws_cl, nr, headers_cl, vals)
        existing_keys.add(key)
        added += 1

# Mark known past outcomes
for r in range(2, ws_cl.max_row + 1):
    ev = (ws_cl.cell(r, 3).value or "").lower()
    cat_d = ws_cl.cell(r, 2).value
    if isinstance(cat_d, datetime):
        cat_d = cat_d.date()
    oc = ws_cl.cell(r, 7)
    if cat_d and cat_d < today and oc.value in ("PENDING", None):
        if "iran sanctions" in ev:
            oc.value = "OCCURRED — waiver not renewed Apr-19; Brent spiked on escalation"
        elif "advance retail" in ev:
            oc.value = "OCCURRED — data released Apr-21"

print(f"CatalystLog: {added} new catalysts added; past outcomes updated")

wb.save("master-data-log.xlsx")
print("Excel sync complete — master-data-log.xlsx saved")
