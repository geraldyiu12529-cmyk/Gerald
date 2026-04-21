"""Append 2026-04-21 trade-rec signals to master-data-log.xlsx."""
from openpyxl import load_workbook
from datetime import date

wb = load_workbook('master-data-log.xlsx')
ws = wb['SignalLedger']
ws2 = wb['AuditAdditionLog']

today = date(2026, 4, 21)
src = '2026-04-21/trade-rec-2026-04-21.md'
regime = 'RISK-ON/GEOPOLITICAL-BINARY-CEASEFIRE-EXPIRED'
vix = 19.04


def append_signal(ID, stype, asset, asset_class, direction, S, T, C, R, Sum_val,
                  entry=None, stop=None, tp1=None, tp2=None,
                  inv=None, inv_date=None, blocking_leg=None, block_reason=None,
                  taken=None, status='OPEN', notes='', gate=''):
    r = ws.max_row + 1
    ws.cell(r, 1, ID)
    ws.cell(r, 2, stype)
    ws.cell(r, 3, today)
    ws.cell(r, 4, asset)
    ws.cell(r, 5, asset_class)
    ws.cell(r, 6, direction)
    ws.cell(r, 7, S)
    ws.cell(r, 8, T)
    ws.cell(r, 9, C)
    ws.cell(r, 10, R)
    ws.cell(r, 11, Sum_val)
    ws.cell(r, 12, entry)
    ws.cell(r, 13, stop)
    ws.cell(r, 14, tp1)
    ws.cell(r, 15, tp2)
    ws.cell(r, 16, inv)
    ws.cell(r, 17, inv_date)
    ws.cell(r, 18, vix)
    ws.cell(r, 19, regime)
    ws.cell(r, 20, blocking_leg)
    ws.cell(r, 21, block_reason)
    ws.cell(r, 22, taken)
    ws.cell(r, 23, status)
    ws.cell(r, 31, src)
    ws.cell(r, 32, notes)
    ws.cell(r, 33, gate)


def append_audit(variable, asset, direction, before, after, impact, decision_moving, source):
    r = ws2.max_row + 1
    ws2.cell(r, 1, today)
    ws2.cell(r, 2, variable)
    ws2.cell(r, 3, asset)
    ws2.cell(r, 4, direction)
    ws2.cell(r, 5, before)
    ws2.cell(r, 6, after)
    ws2.cell(r, 7, impact)
    ws2.cell(r, 8, decision_moving)
    ws2.cell(r, 9, source)


# P010 EWJ
append_signal(
    'P010', 'Promoted', 'EWJ', 'ETF', 'Long', 1, 1, 0, 1, 3,
    entry=90.24, stop=86.00, tp1=95, tp2=98,
    inv='USDJPY<150; Nikkei -5%; BOJ emergency hike >=50bp; US-China trade war escalation',
    inv_date='2026-06-30', taken='PENDING', status='OPEN',
    notes='Promoted cloud-7pm 2026-04-21, re-confirmed local 20:20. Raw TSMOM (ETF). V027 z+1.65 (A). Size 0.75% corr haircut vs SPY.',
    gate='Intl Equity ON')

# P011 Gold gate-blocked
append_signal(
    'P011', 'Promoted', 'Gold', 'Commodity', 'Long', 1, 1, 0, 1, 3,
    inv='Iran deal -> rate hike risk -> gold under pressure; break below $4640',
    inv_date='2026-06-30',
    blocking_leg='Gate',
    block_reason='Commodity sleeve UNCERTAIN (GSCI MISSING 4th run) -> effective OFF; Taken=NO',
    taken='NO', status='GATE-BLOCKED',
    notes='Sum+3 commodity gated. Gold $4808. Gate-blocked at local 20:20.',
    gate='Commodity UNCERTAIN/OFF')

# P012 Silver gate-blocked
append_signal(
    'P012', 'Promoted', 'Silver', 'Commodity', 'Long', 1, 1, 0, 1, 3,
    inv='Iran ceasefire -> commodity deflation; break below $75',
    inv_date='2026-06-30',
    blocking_leg='Gate',
    block_reason='Commodity sleeve UNCERTAIN (GSCI MISSING 4th run) -> effective OFF; Taken=NO',
    taken='NO', status='GATE-BLOCKED',
    notes='Sum+3 commodity gated. Silver $79.07. Gate-blocked at local 20:20.',
    gate='Commodity UNCERTAIN/OFF')

# P013 INTC new promotion
append_signal(
    'P013', 'Promoted', 'INTC', 'Equity', 'Long', 1, 1, 1, 1, 4,
    entry=70.0, stop=65.0, tp1=75, tp2=82,
    inv='Q1 miss on Terafab milestones/gross margin cut -> C=0/-1 -> exit; break below ATR stop; FOMC Apr-28-29 hawkish',
    inv_date='2026-06-23', taken='PENDING', status='OPEN',
    notes='NEW PROMOTION. V026 residual +13.89% (A) T=+1. Sum+4 highest conviction signal. Entry AFTER Apr-23 AC beat (Apr-24 morning). Size 1.0% full (V027 z+1.65 capital expansion). ATR ~$2.50.',
    gate='Equity ON')

# P014 AAPL new promotion
append_signal(
    'P014', 'Promoted', 'AAPL', 'Equity', 'Long', 1, 1, 0, 1, 3,
    entry=273.0, stop=266.5, tp1=280, tp2=290,
    inv='Break below $264 (2xATR); SPY break $696; FOMC hawkish; hard exit 2026-04-30 before May 1 earnings',
    inv_date='2026-04-30', taken='PENDING', status='OPEN',
    notes='NEW PROMOTION. V026 residual +4.68% T=+1. C=0 structural only. Size 0.75% corr haircut vs SPY. Hard time-stop 2026-04-30 BINDING.',
    gate='Equity ON')

# P015 GOOGL contingent
append_signal(
    'P015', 'Promoted', 'GOOGL', 'Equity', 'Long', 1, 0, 1, 1, 3,
    entry=335.0, stop=317.0, tp1=355, tp2=375,
    inv='Earnings miss -> Sum<3 -> no entry; break below 2xATR from entry; FOMC hawkish',
    inv_date='2026-05-15', taken='PENDING', status='CONTINGENT',
    notes='CONTINGENT on Apr-23 UTC+8 beat confirm. V026 residual -0.80% T=0. Sum+3 via S+1/T0/C+1/R+1. Do not chase gap. Size 0.75%.',
    gate='Equity ON')

# N040 MU near-miss
append_signal(
    'N040', 'Near-Miss', 'MU', 'Equity', 'Long', 1, 1, 0, 1, 3,
    blocking_leg='Correlation',
    block_reason='Equity sleeve at capacity (SPY+EWJ+AAPL+INTC) during Apr-22-29 earnings+FOMC cluster',
    taken=None, status='OPEN',
    notes='NEW. V026 residual +15.63% T=+1. Deferred to post-FOMC (Apr-30+). Trigger: position reduction allows entry.',
    gate='Equity ON')

# N041 NVDA near-miss
append_signal(
    'N041', 'Near-Miss', 'NVDA', 'Equity', 'Long', 1, 0, 1, 1, 3,
    blocking_leg='Correlation',
    block_reason='Tech concentration (GOOGL+AAPL+INTC); T=0 (residual -1.00% neutral)',
    taken=None, status='OPEN',
    notes='V026 residual -1.00% T=0 neutral. C+1 AI capex + GOOGL read-through. Trigger: GOOGL beat + residual >+2%.',
    gate='Equity ON')

# AuditAdditionLog entries
append_audit('V026 Residual Momentum', 'INTC', 'T: MISSING->+1 (resid+13.89%)',
             'T=0 blocked', 'T=+1', '+1 Sum -> new Sum+4 PROMOTED P013', 'YES', src)
append_audit('V026 Residual Momentum', 'AAPL', 'T: MISSING->+1 (resid+4.68%)',
             'T=0 blocked', 'T=+1', '+1 Sum -> new Sum+3 PROMOTED P014', 'YES', src)
append_audit('V026 Residual Momentum', 'PLTR', 'T: MISSING->-1 (resid-35.62%)',
             'T=0 blocked', 'T=-1', '-1 Sum -> downgraded to +2 below threshold', 'YES', src)
append_audit('V026 Residual Momentum', 'TSLA', 'T: MISSING->-1 (resid-9.12%)',
             'T=0 blocked', 'T=-1', '-1 Sum -> confirmed no-trade', 'YES', src)
append_audit('V026 Residual Momentum', 'GOOGL', 'T: MISSING->0 (resid-0.80%)',
             'T=0 blocked', 'T=0 resolved', 'Sum+3 confirmed via C+1 (no T change)', 'YES', src)
append_audit('V026 Residual Momentum', 'MU', 'T: MISSING->+1 (resid+15.63%)',
             'T=0 blocked', 'T=+1', 'New Sum+3 near-miss N040', 'YES', src)
append_audit('V027 IC z-score', 'All equity/intl',
             'z: MISSING->+1.65 (fresh Apr-21)',
             'R=+1 proxy-only', 'R=+1 Grade A confirmed',
             'R confidence upgraded; full sizing confirmed per Risk Rules 1.B', 'YES', src)
append_audit('MOVE (Bond Vol Index)', 'All assets',
             '67.90 (A fresh) was MISSING in cloud',
             'R=+1 HY OAS proxy', 'R=+1 Grade A triple (MOVE+HY+V027)',
             'R=+1 now Grade A confirmed across 3 independent inputs', 'YES', src)

wb.save('master-data-log.xlsx')
print('SignalLedger appended: P010 P011 P012 P013 P014 P015 N040 N041')
print('AuditAdditionLog appended: 8 entries')
print('Total SignalLedger rows:', ws.max_row)
print('Total AuditAdditionLog rows:', ws2.max_row)
