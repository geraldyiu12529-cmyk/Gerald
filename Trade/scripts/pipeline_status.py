"""
Pipeline status tracker for Gerald's trading pipeline.

Provides read/write access to .pipeline-status.json with consecutive failure
tracking, task assessment, validity gates, self-heal, and recovery watchdog
management.

Usage by other tasks (paste into task prompts):
    import sys; sys.path.insert(0, './scripts')
    from pipeline_status import PipelineStatus
    ps = PipelineStatus()
    ps.write_ok('market-brief', file='market-brief-2026-04-16.md')
    ps.write_fail('trade-rec', reason='Brief missing', consecutive=True)

Usage by the pipeline-recovery skill (Phase A fast triage):
    ps = PipelineStatus()
    fast = ps.fast_health_check()           # cheap, no Excel, no scripts, no web
    if fast['healthy']:
        ps.clear_watchdog_healthy()
        print(fast['summary'])
        # exit
    ps.start_watchdog()
    # ... Phase B recovery ...
    ps.finish_watchdog(actions, successes, failures)
"""

import json
import os
import re
from pathlib import Path
from datetime import datetime, timezone, timedelta
from typing import Optional


UTC8 = timezone(timedelta(hours=8))

def _resolve_trade_dir() -> Path:
    """Resolve the Trade workspace directory, handling Cowork sandbox mounts."""
    # 1. Explicit env var (already used by catalysts_cache.py)
    if os.environ.get("TRADE_DIR"):
        return Path(os.environ["TRADE_DIR"])
    # 2. Direct mount (native / Windows-side)
    direct = Path('/mnt/Trade')
    if direct.is_dir():
        return direct
    # 3. Cowork sandbox mount pattern: /sessions/*/mnt/Trade
    for p in Path('/sessions').glob('*/mnt/Trade'):
        if p.is_dir():
            return p
    # 4. CWD-relative (Claude Code — skill runs from inside workspace)
    cwd = Path('.')
    if (cwd / 'master-data-log.xlsx').exists():
        return cwd
    # 5. Fallback to the original path (will fail loudly if nothing works)
    return direct

TRADE_DIR = _resolve_trade_dir()
STATUS_PATH = TRADE_DIR / 'pipeline' / '.pipeline-status.json'
HEALTH_PATH = TRADE_DIR / 'pipeline' / '.pipeline-health.json'

# Task schedule (hour, minute) in UTC+8
TASK_SCHEDULE = {
    'preflight':    (19, 45),
    'market-brief': (20, 0),
    'news-events':  (20, 0),
    'trade-rec':    (21, 0),
    'recovery':     (22, 0),
}

# Grace window after scheduled time before a missing file counts as FAILED.
# Covers cron jitter (tasks advertise jitterSeconds up to ~450s ≈ 7.5min) plus
# task runtime. The recovery task fires at 22:00, one hour after trade-rec's
# scheduled 21:00, so even a maximally-jittered + slow trade-rec should be done
# by then. Keep at 20 minutes to stay safe.
GRACE_MINUTES = 20

# Expected output files per task (relative to TRADE_DIR)
TASK_OUTPUTS = {
    'preflight': ['pipeline/.pipeline-health.json', 'audit-data-staging-{date}.md'],
    'market-brief': ['market-brief-{date}.md'],
    'news-events': ['news-events/news-{date}.md'],
    'trade-rec': ['trade-rec-{date}.md', 'report-{date}-trade-rec.html'],
}

# Minimum file sizes (bytes) below which a file is presumed truncated.
# Single source of truth — tasks and recovery both reference this.
MIN_SIZES = {
    'pipeline/.pipeline-health.json': 100,
    'audit-data-staging-{date}.md': 200,
    'market-brief-{date}.md': 1500,
    'news-events/news-{date}.md': 300,
    'trade-rec-{date}.md': 1000,
    'report-{date}-trade-rec.html': 500,
}

# Structural sniff patterns — cheap substring / regex checks run against
# the first SNIFF_READ_BYTES of the file. Empty pattern list means
# "existence + size only".
STRUCTURAL_SNIFFS = {
    'market-brief-{date}.md': [
        r'#\s*(Market\s+Brief|Daily\s+Market\s+Brief)',
        r'Regime\s+Snapshot',   # §1 structural heading — verified 2026-04-14/15/16
        r'Grade',               # evidence-grade rubric referenced
    ],
    'news-events/news-{date}.md': [
        r'#\s*News',
    ],
    'audit-data-staging-{date}.md': [
        r'(residual.?momentum|intermediary|basis.?momentum)',
    ],
    'trade-rec-{date}.md': [
        r'#\s*(Trade\s+Rec|Daily\s+Trade)',
        r'Upstream\s+Synthesis', # §1 structural heading — verified 2026-04-14/15/16
        r'Grade',                # evidence-grade rubric referenced
    ],
}

SNIFF_READ_BYTES = 2048

ESCALATION_THRESHOLD = 3
CRITICAL_THRESHOLD = 5


# ----------------------------------------------------------------------------
# Self-heal helpers for the status JSON
# ----------------------------------------------------------------------------

def _safe_load_status(path: Path) -> tuple:
    """
    Read .pipeline-status.json defensively.

    Returns (status_dict, load_note). load_note is empty on clean read,
    or describes the corruption/self-heal if the file was unreadable.
    """
    if not path.exists():
        return {}, ''
    try:
        text = path.read_text()
    except OSError as e:
        return {}, f'status file unreadable: {e!r}'
    if not text.strip():
        return {}, 'status file was empty'
    try:
        data = json.loads(text)
        if not isinstance(data, dict):
            return {}, f'status file had unexpected type {type(data).__name__}'
        return data, ''
    except json.JSONDecodeError as e:
        # Self-heal: move the corrupt file aside so it's preserved for forensics
        # and return an empty dict. Recovery rebuilds today's view from file
        # evidence; prior-day history is intentionally not fabricated.
        try:
            backup = path.with_suffix(path.suffix + '.corrupt')
            path.rename(backup)
            return {}, f'status file corrupt ({e.msg}); moved to {backup.name}'
        except OSError as move_err:
            return {}, f'status file corrupt and could not be moved aside: {move_err!r}'


# ----------------------------------------------------------------------------
# File validity gate (existence + size + structural sniff)
# ----------------------------------------------------------------------------

def _file_valid(path: Path, min_size: int, patterns: list) -> tuple:
    """
    Check a single output file: exists, size ≥ min, structural sniff passes.
    Returns (ok, reason_if_not_ok).
    """
    if not path.exists():
        return False, 'missing'
    try:
        size = path.stat().st_size
    except OSError as e:
        return False, f'stat failed: {e!r}'
    if size < min_size:
        return False, f'truncated ({size} < {min_size} bytes)'
    if patterns:
        try:
            head = path.read_bytes()[:SNIFF_READ_BYTES].decode('utf-8', errors='replace')
        except OSError as e:
            return False, f'unreadable: {e!r}'
        for pat in patterns:
            if not re.search(pat, head, re.IGNORECASE):
                return False, f'structural sniff failed (no match for /{pat}/i)'
    return True, ''


# ----------------------------------------------------------------------------
# PipelineStatus class
# ----------------------------------------------------------------------------

class PipelineStatus:
    def __init__(self, trade_dir: Optional[Path] = None):
        self.trade_dir = trade_dir or TRADE_DIR
        self.status_path = self.trade_dir / 'pipeline' / '.pipeline-status.json'
        self.now = datetime.now(UTC8)
        self.today = self.now.strftime('%Y-%m-%d')
        self.status, self.load_note = _safe_load_status(self.status_path)
        self._status_was_self_healed = 'corrupt' in (self.load_note or '')

    # ---- I/O ----

    def _write(self):
        """Atomic write via tmp-file + rename."""
        tmp = self.status_path.with_suffix(self.status_path.suffix + '.tmp')
        tmp.write_text(json.dumps(self.status, indent=2))
        tmp.replace(self.status_path)

    def _get_prev(self, task: str) -> dict:
        return self.status.get(task, {})

    def _calc_consecutive(self, task: str, is_failure: bool) -> int:
        if not is_failure:
            return 0
        prev = self._get_prev(task)
        if prev.get('status') in ('OK', None):
            return 1
        return prev.get('consecutive_failures', 0) + 1

    # ---- Writers (called by each task on completion) ----

    def write_ok(self, task: str, file: str = '', missing_count: int = 0,
                 details: str = ''):
        """Task completed successfully."""
        entry = {
            'date': self.today,
            'status': 'PARTIAL' if missing_count > 3 else 'OK',
            'consecutive_failures': 0,
            'timestamp': self.now.isoformat(),
        }
        if file:
            entry['file'] = file
        if missing_count:
            entry['missing_count'] = missing_count
        if details:
            entry['details'] = details
        self.status[task] = entry
        self._write()

    def write_fail(self, task: str, reason: str = '', consecutive: bool = True):
        """Task failed or aborted."""
        consec = self._calc_consecutive(task, True) if consecutive else 0
        self.status[task] = {
            'date': self.today,
            'status': 'FAIL',
            'consecutive_failures': consec,
            'timestamp': self.now.isoformat(),
            'reason': reason,
        }
        self._write()

    def write_abort(self, task: str, reason: str = ''):
        """Task aborted due to upstream integrity failure."""
        consec = self._calc_consecutive(task, True)
        self.status[task] = {
            'date': self.today,
            'status': 'ABORT',
            'consecutive_failures': consec,
            'timestamp': self.now.isoformat(),
            'reason': reason,
        }
        self._write()

    def write_recovery(self, task: str, details: str = ''):
        """Mark a task as recovered (sets PARTIAL, resets streak)."""
        self.status[task] = {
            'date': self.today,
            'status': 'PARTIAL',
            'consecutive_failures': 0,
            'timestamp': self.now.isoformat(),
            'recovery': True,
            'details': details,
        }
        self._write()

    # ---- Recovery task watchdog ----

    def start_watchdog(self):
        """
        Mark recovery as in-progress. Called at the top of Phase B.
        If the recovery task crashes before finish_watchdog(), next morning's
        preflight will see in_progress=True with a non-today date and flag it.
        """
        entry = self.status.get('recovery', {})
        entry.update({
            'date': self.today,
            'in_progress': True,
            'started': self.now.isoformat(),
        })
        # Clear any stale completion fields from a prior run
        for k in ('actions_taken', 'recoveries_succeeded', 'recoveries_failed',
                  'finished', 'status', 'notes', 'healthy_short_circuit'):
            entry.pop(k, None)
        self.status['recovery'] = entry
        self._write()

    def finish_watchdog(self, actions: list, successes: int, failures: int,
                        healthy: bool = False, notes: str = ''):
        """Clear in_progress and write the recovery summary."""
        entry = {
            'date': self.today,
            'timestamp': self.now.isoformat(),
            'finished': self.now.isoformat(),
            'in_progress': False,
            'status': 'OK' if (failures == 0 and not str(notes).startswith('FAIL')) else 'PARTIAL',
            'actions_taken': actions,
            'recoveries_succeeded': successes,
            'recoveries_failed': failures,
            'healthy_short_circuit': healthy,
        }
        if notes:
            entry['notes'] = notes
        self.status['recovery'] = entry
        self._write()

    def clear_watchdog_healthy(self):
        """Shortcut for the cheap healthy-path exit."""
        self.finish_watchdog(actions=[], successes=0, failures=0,
                             healthy=True, notes='')

    def detect_prior_recovery_crash(self) -> Optional[dict]:
        """
        Called by preflight (or any task) to check if a prior recovery task
        died mid-run. Returns the stale entry if so, else None.
        """
        rec = self.status.get('recovery', {})
        if rec.get('in_progress') and rec.get('date') != self.today:
            return rec
        return None

    # ---- Readers ----

    def assess_pipeline(self) -> dict:
        """
        Assess the state of each pipeline task for today.
        Returns dict of task -> {state, consecutive_failures, files, status_entry}.
        State values: HEALTHY | PARTIAL | FAILED | NOT_YET.
        """
        assessment = {}
        for task, (sched_h, sched_m) in TASK_SCHEDULE.items():
            if task == 'recovery':
                continue

            entry = self.status.get(task, {})
            evidence = self._check_files(task)
            past_schedule = self._past_schedule(sched_h, sched_m)

            if entry.get('date') == self.today and entry.get('status') == 'OK':
                state = 'HEALTHY' if evidence['all_valid'] else 'PARTIAL'
            elif entry.get('date') == self.today and entry.get('status') == 'PARTIAL':
                state = 'PARTIAL'
            elif entry.get('date') == self.today and entry.get('status') in ('FAIL', 'ABORT'):
                state = 'FAILED'
            elif entry.get('date') == self.today and entry.get('status') is None:
                state = 'PARTIAL' if evidence['all_valid'] else 'FAILED'
            elif not past_schedule:
                state = 'NOT_YET'
            elif evidence['all_valid'] and not entry:
                # Files valid but no status entry — task completed but crashed
                # before writing status. Treat as healthy.
                state = 'HEALTHY'
            else:
                state = 'FAILED'

            assessment[task] = {
                'state': state,
                'consecutive_failures': entry.get('consecutive_failures', 0),
                'files_valid': evidence['valid'],
                'files_invalid': evidence['invalid'],
                'all_files_valid': evidence['all_valid'],
                'status_entry': entry,
            }

        return assessment

    def _past_schedule(self, sched_h: int, sched_m: int) -> bool:
        """Is now past (sched + grace)?"""
        return (
            self.now.hour > sched_h or
            (self.now.hour == sched_h and self.now.minute > sched_m + GRACE_MINUTES)
        )

    def _check_files(self, task: str) -> dict:
        """Per-task file validity check using MIN_SIZES and STRUCTURAL_SNIFFS."""
        templates = TASK_OUTPUTS.get(task, [])
        valid = []
        invalid = []
        for tmpl in templates:
            path = self.trade_dir / tmpl.format(date=self.today)
            min_size = MIN_SIZES.get(tmpl, 1)
            patterns = STRUCTURAL_SNIFFS.get(tmpl, [])
            ok, reason = _file_valid(path, min_size, patterns)
            if ok:
                valid.append(path.name)
            else:
                invalid.append({'name': path.name, 'reason': reason})
        return {
            'valid': valid,
            'invalid': invalid,
            'all_valid': len(invalid) == 0,
        }

    def fast_health_check(self) -> dict:
        """
        Cheap triage pass. No Excel, no script execution, no web.
        Returns:
            {
                'healthy': bool,
                'assessment': {task: {state, ...}},
                'escalations': [...],
                'self_healed': bool,
                'load_note': str,
                'prior_recovery_crash': dict | None,
                'summary': str
            }
        """
        assessment = self.assess_pipeline()
        escalations = self.get_escalations()

        all_good_states = all(
            info['state'] in ('HEALTHY', 'NOT_YET')
            for info in assessment.values()
        )
        no_critical = all(e['level'] != 'CRITICAL' for e in escalations)
        prior_crash = self.detect_prior_recovery_crash()
        healthy = (
            all_good_states and no_critical and
            not self._status_was_self_healed and
            prior_crash is None
        )

        states = ', '.join(
            f"{task}={info['state']}" for task, info in assessment.items()
        )
        summary = f"Pipeline {'healthy' if healthy else 'needs attention'} — {states}"
        if escalations:
            summary += f" | escalations: {len(escalations)}"
        if self._status_was_self_healed:
            summary += f" | status JSON self-healed: {self.load_note}"
        if prior_crash:
            summary += f" | prior recovery crash on {prior_crash.get('date')}"

        return {
            'healthy': healthy,
            'assessment': assessment,
            'escalations': escalations,
            'self_healed': self._status_was_self_healed,
            'load_note': self.load_note,
            'prior_recovery_crash': prior_crash,
            'summary': summary,
        }

    def get_escalations(self) -> list:
        """Return tasks that need escalation based on consecutive failures."""
        escalations = []
        for task in TASK_SCHEDULE:
            if task == 'recovery':
                continue
            entry = self.status.get(task, {})
            consec = entry.get('consecutive_failures', 0)
            if consec >= CRITICAL_THRESHOLD:
                escalations.append({
                    'task': task,
                    'level': 'CRITICAL',
                    'consecutive': consec,
                    'action': 'Write to framework/Memory.md',
                })
            elif consec >= ESCALATION_THRESHOLD:
                escalations.append({
                    'task': task,
                    'level': 'WARNING',
                    'consecutive': consec,
                    'action': 'Flag in output',
                })
        return escalations

    def is_pipeline_healthy(self) -> bool:
        """Legacy helper — kept for existing callers. Prefer fast_health_check()."""
        return self.fast_health_check()['healthy']

    def format_summary(self, verbose: bool = False) -> str:
        """Human-readable pipeline status summary."""
        fast = self.fast_health_check()
        assessment = fast['assessment']
        escalations = fast['escalations']

        lines = [f"## Pipeline Status — {self.today}\n"]

        for task, info in assessment.items():
            state = info['state']
            consec = info['consecutive_failures']
            icon = {'HEALTHY': 'OK', 'PARTIAL': 'PARTIAL', 'FAILED': 'FAIL',
                    'NOT_YET': 'PENDING'}[state]
            line = f"- **{task}**: {icon}"
            if consec > 0:
                line += f" (consecutive failures: {consec})"
            if info['files_invalid'] and verbose:
                probs = ', '.join(
                    f"{f['name']}={f['reason']}" for f in info['files_invalid']
                )
                line += f" — issues: {probs}"
            lines.append(line)

        if escalations:
            lines.append("\n### Escalations")
            for esc in escalations:
                lines.append(
                    f"- **{esc['level']}**: `{esc['task']}` — "
                    f"{esc['consecutive']} consecutive failures. "
                    f"Action: {esc['action']}"
                )

        if fast['self_healed']:
            lines.append(f"\n### Notes\n- Status JSON self-healed: {fast['load_note']}")
        if fast['prior_recovery_crash']:
            lines.append(
                f"- Prior recovery crash detected on "
                f"{fast['prior_recovery_crash'].get('date')} — investigate"
            )

        return '\n'.join(lines)


# ----------------------------------------------------------------------------
# Excel sync integrity (Phase B only — skipped on healthy days)
# ----------------------------------------------------------------------------

def excel_sync_check(xlsx_path: Optional[Path] = None,
                     today: Optional[str] = None) -> dict:
    """
    Read-only check that master-data-log.xlsx was updated today.
    Opens xlsx via openpyxl in read-only mode. Never raises — converts any
    error to a failure entry. Gerald having the xlsx open in Excel is a
    soft failure (we flag it as a possibility, not a hard error).
    """
    if today is None:
        today = datetime.now(UTC8).strftime('%Y-%m-%d')
    xlsx_path = xlsx_path or (TRADE_DIR / 'master-data-log.xlsx')

    report = {
        'path': str(xlsx_path),
        'exists': xlsx_path.exists(),
        'checks': {},
        'ok': False,
        'note': '',
    }
    if not xlsx_path.exists():
        report['note'] = 'master-data-log.xlsx not found'
        return report

    try:
        from openpyxl import load_workbook
    except ImportError:
        report['note'] = 'openpyxl unavailable; skipping Excel integrity check'
        return report

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        report['note'] = f'could not open xlsx (may be locked by Excel): {type(e).__name__}'
        return report

    def _last_non_empty_col_a(sheet):
        last = None
        for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
            if row[0]:
                last = row[0]
        if isinstance(last, datetime):
            return last.strftime('%Y-%m-%d')
        return str(last) if last is not None else None

    for sheet_name in ('DailyVariables', 'RegimeHistory'):
        try:
            if sheet_name in wb.sheetnames:
                last = _last_non_empty_col_a(wb[sheet_name])
                report['checks'][f'{sheet_name}_last_date'] = last
                report['checks'][f'{sheet_name}_is_today'] = (last == today)
            else:
                report['checks'][sheet_name] = 'sheet missing'
        except Exception as e:
            report['checks'][f'{sheet_name}_error'] = f'{type(e).__name__}: {e}'

    try:
        wb.close()
    except Exception:
        pass

    report['ok'] = bool(
        report['checks'].get('DailyVariables_is_today') and
        report['checks'].get('RegimeHistory_is_today')
    )
    return report


# ----------------------------------------------------------------------------
# Recovery-brief structural validator
# ----------------------------------------------------------------------------

def validate_recovery_brief(path: Path) -> tuple:
    """
    Post-write check on a recovery brief. Returns (ok, missing_sections).
    The trade rec consumes: header, regime label, and at least one asset
    scorecard line with S/T/C/R columns. A recovery brief must also carry
    a RECOVERY disclosure so nobody mistakes it for a full brief.
    """
    if not path.exists():
        return False, ['file missing']
    try:
        text = path.read_text()
    except OSError as e:
        return False, [f'read failed: {e!r}']

    missing = []
    if not re.search(r'#\s*Market\s+Brief', text, re.IGNORECASE):
        missing.append('brief header')
    if not re.search(r'regime', text, re.IGNORECASE):
        missing.append('regime label')
    if not re.search(r'\bS\s*\|\s*T\s*\|\s*C\s*\|\s*R\b', text):
        missing.append('S|T|C|R scorecard')
    if not re.search(r'recovery', text, re.IGNORECASE):
        missing.append('RECOVERY disclosure')
    return (len(missing) == 0), missing


# ----------------------------------------------------------------------------
# Convenience functions for task prompt integration (unchanged public API)
# ----------------------------------------------------------------------------

def check_upstream_for_trade_rec() -> tuple:
    """
    Pre-check for the trade rec task. Returns (can_proceed, message).
    Hard-aborts if brief is missing OR truncated OR structurally broken.
    """
    ps = PipelineStatus()
    today = ps.today
    brief = TRADE_DIR / f'market-brief-{today}.md'
    news = TRADE_DIR / f'news-events/news-{today}.md'
    staging = TRADE_DIR / f'audit-data-staging-{today}.md'

    brief_ok, brief_reason = _file_valid(
        brief, MIN_SIZES['market-brief-{date}.md'],
        STRUCTURAL_SNIFFS['market-brief-{date}.md']
    )
    news_ok, news_reason = _file_valid(
        news, MIN_SIZES['news-events/news-{date}.md'],
        STRUCTURAL_SNIFFS['news-events/news-{date}.md']
    )
    staging_ok, staging_reason = _file_valid(
        staging, MIN_SIZES['audit-data-staging-{date}.md'],
        STRUCTURAL_SNIFFS['audit-data-staging-{date}.md']
    )

    missing = []
    if not brief_ok:   missing.append(f'market-brief-{today}.md ({brief_reason})')
    if not news_ok:    missing.append(f'news-events/news-{today}.md ({news_reason})')
    if not staging_ok: missing.append(f'audit-data-staging-{today}.md ({staging_reason})')

    if not brief_ok:
        ps.write_abort('trade-rec', reason=f"Missing/invalid: {', '.join(missing)}")
        consec = ps.status.get('trade-rec', {}).get('consecutive_failures', 0)
        return False, (
            f"PIPELINE INTEGRITY FAIL — {', '.join(missing)}. "
            f"Trade rec aborted. Consecutive failures: {consec}."
        )

    if missing:
        return True, (
            f"PIPELINE INTEGRITY WARNING — {', '.join(missing)}. "
            f"Proceeding with partial data."
        )

    return True, "Upstream integrity check passed."


def check_briefs_for_weekly_review() -> tuple:
    """
    Pre-check for weekly regime/signal review.
    Needs 4/5 valid weekday briefs from the most recent trading week.
    """
    ps = PipelineStatus()
    now = datetime.now(UTC8)

    days_back = (now.weekday() + 2) % 7  # days since last Friday
    if days_back == 0:
        days_back = 7
    last_friday = now - timedelta(days=days_back)
    week_dates = []
    for i in range(5):
        d = last_friday - timedelta(days=4 - i)
        week_dates.append(d.strftime('%Y-%m-%d'))

    found = 0
    missing_dates = []
    for date_str in week_dates:
        brief = TRADE_DIR / f'market-brief-{date_str}.md'
        ok, _ = _file_valid(
            brief, MIN_SIZES['market-brief-{date}.md'],
            STRUCTURAL_SNIFFS['market-brief-{date}.md']
        )
        if ok:
            found += 1
        else:
            missing_dates.append(date_str)

    if found <= 1:
        return False, (
            f"PIPELINE INTEGRITY FAIL — only {found}/5 valid weekly briefs found "
            f"(missing/invalid: {', '.join(missing_dates)}). "
            f"Insufficient data for regime aggregation. Weekly review aborted."
        )
    elif found < 4:
        return True, (
            f"PIPELINE INTEGRITY WARNING — only {found}/5 valid weekly briefs "
            f"(missing/invalid: {', '.join(missing_dates)}). "
            f"Regime aggregation may be incomplete. Proceeding with low confidence."
        )
    else:
        return True, f"Weekly brief coverage: {found}/5 days valid."


# ----------------------------------------------------------------------------
# CLI entry point for ad-hoc inspection and for the scheduled task's fast path
# ----------------------------------------------------------------------------

if __name__ == '__main__':
    import sys
    ps = PipelineStatus()
    mode = sys.argv[1] if len(sys.argv) > 1 else 'summary'
    if mode == 'fast':
        fast = ps.fast_health_check()
        print(fast['summary'])
        sys.exit(0 if fast['healthy'] else 2)
    elif mode == 'verbose':
        print(ps.format_summary(verbose=True))
    elif mode == 'json':
        fast = ps.fast_health_check()
        print(json.dumps(fast, indent=2, default=str))
    else:
        print(ps.format_summary(verbose=False))
