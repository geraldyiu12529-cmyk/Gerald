"""
sync_trade_exits_2026-04-20.py
Single-use script: marks P003 (INTC), P004 (Gold), P005 (QQQ) as closed
in the SignalLedger sheet of master-data-log.xlsx.
Run once from the Trade/ directory.
"""

from openpyxl import load_workbook
from datetime import date

WB_PATH = "C:/Users/Lokis/OneDrive/Desktop/T.system/claude-code/Trade/master-data-log.xlsx"

wb = load_workbook(WB_PATH)
ws = wb["SignalLedger"]

# Read header row to build column index
headers = {cell.value: cell.column for cell in ws[1] if cell.value}
print("Columns found:", list(headers.keys()))

# Define updates keyed by ID (column A = ID)
updates = {
    "P003": {
        "Status":           "HIT_TARGET",
        "Exit_Price":       68.26,
        "Exit_Date":        date(2026, 4, 19),
        "Days_to_Exit":     3,
        "Hypo_PnL_Pct":    0.57,   # (68.26/67.87 - 1) * 100
        "Catalyst_Outcome": "EXITED_PRE_CATALYST",
        "Notes":            "Closed 2026-04-19 21:31 UTC+8 before Apr-23 earnings. "
                            "Memory.md entry $64.68 vs Binance fill $67.87 — entry-price discrepancy noted.",
    },
    "P004": {
        "Status":           "HIT_TARGET",
        "Exit_Price":       4823.65,
        "Exit_Date":        date(2026, 4, 18),
        "Days_to_Exit":     2,
        "Hypo_PnL_Pct":    0.90,   # (4823.65/4780.69 - 1) * 100
        "Catalyst_Outcome": "EXITED_PRE_CATALYST",
        "Notes":            "Closed 2026-04-18 17:12 UTC+8 before Apr-22 ceasefire binary. "
                            "DXY + real-yield thesis intact at exit.",
    },
    "P005": {
        "Status":           "HIT_TARGET",
        "Exit_Price":       643.02,
        "Exit_Date":        date(2026, 4, 19),
        "Days_to_Exit":     2,
        "Hypo_PnL_Pct":    0.46,   # (643.02/640.09 - 1) * 100
        "Catalyst_Outcome": "EXITED_PRE_CATALYST",
        "Notes":            "Closed 2026-04-19 16:00 UTC+8 before Apr-22–23 earnings cluster. "
                            "FOMC Apr-28 unresolved at exit.",
    },
}

updated = []
for row in ws.iter_rows(min_row=2):
    row_id = row[0].value
    if row_id in updates:
        for field, value in updates[row_id].items():
            if field in headers:
                ws.cell(row=row[0].row, column=headers[field], value=value)
            else:
                print(f"  WARNING: column '{field}' not found — skipping for {row_id}")
        updated.append(row_id)
        print(f"  Updated {row_id}: Status=HIT_TARGET, Exit_Date={updates[row_id]['Exit_Date']}")

if len(updated) == 3:
    wb.save(WB_PATH)
    print(f"\nSaved. Updated rows: {updated}")
else:
    print(f"\nWARNING: Expected 3 updates, got {len(updated)}. IDs found: {updated}. File NOT saved.")
